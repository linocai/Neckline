"""交割单解析(plan 4D.1)。输入券商导出的 xlsx(bytes)→ 规范化 `RawTrade` 列表。

**两家券商 schema(用户样例文件已不在磁盘,以下结构从真实文件逐字段核实过,任务
指令原文钉死,不臆造)**:

    · 券商一格式(手机截图整理,无成交价,也【无证券代码列】):
      交易日期(datetime) / 券商来源 / 业务名称 / 证券名称 / 成交数量 / 股份余额 /
      费用 / 发生金额 / 资金余额 / 备注。业务名称枚举:证券买入清算、证券卖出清算、
      银证转入、银证转出、红股派息、股息红利个人所得税扣款、利息归本。
      成交价反推:`(|发生金额| - 费用) / 数量`(公式按任务指令原样实现,不反向验证
      买卖方向下的正负号语义——发生金额/费用的符号约定由该券商决定,本模块不猜)。
      **无代码列 → ts_code 必须靠"证券名称"反查**(`resolve_code_by_name`)。

    · 券商二格式(桌面导出):交易日期 / 券商来源 / 证券代码 / 证券名称 / 业务名称 /
      成交价格 / 成交数量 / 成交金额 / 佣金 / 印花税 / 过户费 / 清算费(B股) /
      发生金额 / 资金余额 / 股份余额 / 股东代码 / 币种 / 合同编号。业务名称枚举:
      证券买入、证券卖出、银行转证券、指定交易。**证券代码等字段可能带零宽空格
      (U+200B)前缀,解析时 strip**;代码无交易所后缀,按 6/0/3/8-4-920 前缀补
      .SH/.SZ/.BJ(复用 `neckline.sentinel.quotes.to_symbol`,它本身已复用
      `neckline.data.board.classify_by_code` 单一源,本模块不再另写一份前缀正则)。

通用清洗:跳过非交易行(指定交易 / 银证转账等资金流水,非成交)、金额正负号由
该券商自行约定(反推公式照抄任务指令,不做符号假设)、按(代码,日期序)供上层
FIFO 闭合回合(本模块只产出规范化交易流,不做回合闭合——见 `reconcile.py`)。

**可配字段映射(plan 4D.1「留 review_col_map 可覆盖以支持两家券商原始格式」)**:
`col_map` 是「规范字段 → 该工作簿实际列名」的覆盖字典,规范字段固定为 9 个
(`CANONICAL_FIELDS`);未覆盖的字段退化到 `_DEFAULT_HEADERS` 的内置默认列名。
表头行探测固定按"含《交易日期》列"(任务指令原文),不受 col_map 影响——col_map
只remap 其它字段的列名,不改表头探测锚点。
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional

import polars as pl

from neckline.data.market_data import load_namechange, load_stock_basic
from neckline.sentinel.quotes import to_symbol

# —— 规范字段(plan 4D.1 原文枚举,col_map 的 key 空间)——————————————————————
CANONICAL_FIELDS = (
    "成交日期", "代码", "名称", "买卖方向", "成交价",
    "成交数量", "成交金额", "手续费", "印花税", "过户费",
)

# —— 业务名称枚举(两家券商固定值,非 col_map 可配范围——col_map 只 remap 列名,
#    不 remap 列值语义)——————————————————————————————————————————————————
_BUY_NAMES = {"format1": {"证券买入清算"}, "format2": {"证券买入"}}
_SELL_NAMES = {"format1": {"证券卖出清算"}, "format2": {"证券卖出"}}
_SKIP_NAMES = {
    "format1": {"银证转入", "银证转出", "红股派息", "股息红利个人所得税扣款", "利息归本"},
    "format2": {"银行转证券", "指定交易"},
}

# —— 内置默认列名(未被 col_map 覆盖时的兜底;§4D.1"先支持整理格式,留可配支持
#    两家券商原始格式")—————————————————————————————————————————————————
_DEFAULT_HEADERS = {
    "format1": {
        "成交日期": "交易日期",
        "名称": "证券名称",
        "买卖方向": "业务名称",
        "成交数量": "成交数量",
        "成交金额": "发生金额",   # 格式一无独立"成交金额"列,兜底用发生金额(反推价格用)
        "手续费": "费用",
    },
    "format2": {
        "成交日期": "交易日期",
        "代码": "证券代码",
        "名称": "证券名称",
        "买卖方向": "业务名称",
        "成交价": "成交价格",
        "成交数量": "成交数量",
        "成交金额": "成交金额",
        "手续费": "佣金",
        "印花税": "印花税",
        "过户费": "过户费",
    },
}
_FORMAT2_CLEARING_FEE_HEADER = "清算费(B股)"   # 不入 col_map(非任务指令枚举的 9 个规范字段之一),固定按此列名探测,缺失视为 0

_HEADER_SCAN_ROWS = 15   # 表头前最多容许 1-3 行标题/说明,留宽松缓冲

# 零宽字符(U+200B 零宽空格为任务指令明确提到的坑;顺带清一批同类不可见字符防御)。
_INVISIBLE_CHARS_RE = re.compile("[​‌‍﻿]")


def clean_str(v: object) -> str:
    """去首尾空白 + 零宽字符(§4D.1「证券代码等字段可能带零宽空格前缀」)。"""
    if v is None:
        return ""
    s = str(v)
    s = _INVISIBLE_CHARS_RE.sub("", s)
    s = unicodedata.normalize("NFKC", s)
    return s.strip()


def normalize_ts_code(raw: str) -> str:
    """裸 6 位代码(可能带零宽空格)→ 带交易所后缀的 ts_code。**不写正则**——
    复用 `neckline.sentinel.quotes.to_symbol`(已复用 `board.classify_by_code`
    单一源判定北交所前缀,§4D.1「勿另写正则」)。已带 `.SH/.SZ/.BJ` 后缀的原样
    规范化大写返回。"""
    s = clean_str(raw)
    if not s:
        return s
    if re.match(r"^\d{6}\.[A-Za-z]{2}$", s):
        digits, suffix = s.split(".")
        return f"{digits}.{suffix.upper()}"
    digits = re.sub(r"\D", "", s)
    if not digits:
        return s
    sym = to_symbol(digits)          # 如 "sh600519" / "sz000001" / "bj920117"
    prefix, code_digits = sym[:2], sym[2:]
    suffix = {"sh": "SH", "sz": "SZ", "bj": "BJ"}.get(prefix, "SH")
    return f"{code_digits}.{suffix}"


def _parse_cell_date(v: object) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = clean_str(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_cell_float(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = clean_str(v).replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


@dataclass
class NameIndex:
    """"证券名称 → ts_code" 反查索引(供格式一——无代码列——用)。按 trade_date
    做 as-of 区间匹配,`namechange` 覆盖不到的(已知有分页缺行小概率,阶段0 遗留
    问题 4)回退 `stock_basic` 当前名称精确匹配。同名映射到 >1 个不同 ts_code
    (歧义)时诚实返回 None,不瞎猜。"""

    namechange: pl.DataFrame
    stock_basic: pl.DataFrame

    def resolve(self, name: str, trade_date: date) -> Optional[str]:
        name = clean_str(name)
        if not name:
            return None
        codes: set = set()
        if not self.namechange.is_empty():
            sub = self.namechange.filter(pl.col("name") == name)
            for row in sub.iter_rows(named=True):
                start = row.get("start_date")
                end = row.get("end_date")
                if start is not None and start > trade_date:
                    continue
                if end is not None and end < trade_date:
                    continue
                codes.add(row["ts_code"])
        if len(codes) == 1:
            return next(iter(codes))
        if len(codes) > 1:
            return None  # 歧义:同名在同一时点映射到多个代码,不猜
        if not self.stock_basic.is_empty():
            sub2 = self.stock_basic.filter(pl.col("name") == name)
            codes2 = set(sub2["ts_code"].to_list())
            if len(codes2) == 1:
                return next(iter(codes2))
        return None


def build_name_index(db_path: Optional[Path] = None) -> NameIndex:
    return NameIndex(namechange=load_namechange(db_path), stock_basic=load_stock_basic(db_path))


@dataclass
class RawTrade:
    trade_date: date
    ts_code: str
    name: str
    side: str                 # "buy" | "sell"
    price: float               # 成交价(格式一反推 / 格式二直读)
    qty: int                   # 成交数量(股,正数)
    fee: float                 # 本笔归属费用(绝对值;格式一=|费用|,格式二=佣金+印花税+过户费+清算费(B股) 各自取绝对值求和)
    cash_flow: float            # 发生金额(原始符号,供审计留痕,不参与三查计算)
    broker_source: str = ""     # 券商来源(如有)
    source_format: str = ""     # "format1" | "format2"
    source_ref: str = ""        # "<sheet>!row<N>",诊断用


@dataclass
class ParseWarning:
    sheet: str
    row: int
    message: str


@dataclass
class ParseResult:
    trades: List[RawTrade] = field(default_factory=list)
    warnings: List[ParseWarning] = field(default_factory=list)
    sheet_formats: Dict[str, str] = field(default_factory=dict)   # sheet名 -> "format1"/"format2"/"skipped:<reason>"
    skip_counts: Dict[str, int] = field(default_factory=dict)     # 业务名称(非成交行)-> 跳过行数,供展示留痕


def _find_header_row(rows: List[List[object]]) -> Optional[int]:
    """在前 `_HEADER_SCAN_ROWS` 行内找"含《交易日期》列"的那一行(任务指令原文
    锚点,不受 col_map 影响)。返回该行在 `rows` 中的下标,找不到 → None。"""
    limit = min(len(rows), _HEADER_SCAN_ROWS)
    for i in range(limit):
        cells = [clean_str(c) for c in rows[i]]
        if "交易日期" in cells:
            return i
    return None


def _detect_format(headers: List[str], col_map: Optional[Dict[str, str]] = None) -> Optional[str]:
    """格式二有独立"证券代码"列(格式一没有,只有"证券名称")——用它作主判据;
    "成交价格"列是次要判据(格式一无成交价,靠反推)。两者皆无 → 判格式一
    (格式一字段集是格式二的子集去掉"证券代码"/"成交价格"等,故默认落这里)。

    判据列名本身也吃 col_map 覆盖(若用户的券商格式把"证券代码"/"成交价格"/
    "证券名称"这几个判据列改了名,不 remap 判据就永远探测不出格式,col_map 就
    形同虚设)——表头行探测锚点("交易日期")仍固定不受影响,这里只是"认出
    是哪种格式"的判据列名可配。"""
    m = col_map or {}
    code_col = m.get("代码", "证券代码")
    price_col = m.get("成交价", "成交价格")
    name_col = m.get("名称", "证券名称")
    if code_col in headers or price_col in headers:
        return "format2"
    if name_col in headers:
        return "format1"
    return None


def _col_index(headers: List[str], canonical: str, fmt: str, col_map: Optional[Dict[str, str]]) -> Optional[int]:
    override = (col_map or {}).get(canonical)
    target = override if override else _DEFAULT_HEADERS[fmt].get(canonical)
    if not target:
        return None
    try:
        return headers.index(target)
    except ValueError:
        return None


def _cell(row: List[object], idx: Optional[int]) -> object:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parse_workbook(
    data: bytes,
    filename: str = "",
    *,
    col_map: Optional[Dict[str, str]] = None,
    db_path: Optional[Path] = None,
    name_index: Optional[NameIndex] = None,
) -> ParseResult:
    """解析一份交割单 xlsx(多 sheet)。任何单行解析异常降级为该行的一条
    `ParseWarning`(跳过该行,不中断整份解析);整份工作簿打不开(非法 xlsx)→
    抛 `ValueError`,由 API 层转 400(这是"文件本身不是 xlsx"的用户可纠正错误,
    与"某一行数据有瑕疵"的优雅降级是两回事)。"""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"{filename or '文件'} 不是合法的 xlsx 工作簿:{e}") from e

    result = ParseResult()
    idx = name_index if name_index is not None else build_name_index(db_path)

    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            result.sheet_formats[ws.title] = "skipped:empty"
            continue
        header_i = _find_header_row(rows)
        if header_i is None:
            result.sheet_formats[ws.title] = "skipped:no_header(未找到含'交易日期'的表头行)"
            continue
        headers = [clean_str(c) for c in rows[header_i]]
        fmt = _detect_format(headers, col_map)
        if fmt is None:
            result.sheet_formats[ws.title] = "skipped:unknown_format"
            result.warnings.append(ParseWarning(ws.title, header_i + 1, "表头列不匹配已知的两种券商格式,已跳过整个 sheet。"))
            continue
        result.sheet_formats[ws.title] = fmt

        cols = {c: _col_index(headers, c, fmt, col_map) for c in CANONICAL_FIELDS}
        clearing_fee_idx = headers.index(_FORMAT2_CLEARING_FEE_HEADER) if _FORMAT2_CLEARING_FEE_HEADER in headers else None

        for r in range(header_i + 1, len(rows)):
            row = rows[r]
            if row is None or all(c is None or clean_str(c) == "" for c in row):
                continue
            ref = f"{ws.title}!row{r + 1}"
            try:
                trade, skip_label, warn_msg = _parse_row(
                    row, fmt, cols, clearing_fee_idx, idx, ref,
                )
            except Exception as e:  # noqa: BLE001  单行异常绝不中断整份解析
                result.warnings.append(ParseWarning(ws.title, r + 1, f"解析异常已跳过该行:{e}"))
                continue
            if warn_msg:
                result.warnings.append(ParseWarning(ws.title, r + 1, warn_msg))
            if skip_label:
                result.skip_counts[skip_label] = result.skip_counts.get(skip_label, 0) + 1
            if trade is not None:
                result.trades.append(trade)

    return result


def _parse_row(
    row: List[object],
    fmt: str,
    cols: Dict[str, Optional[int]],
    clearing_fee_idx: Optional[int],
    name_index: NameIndex,
    ref: str,
):
    """返回 (RawTrade|None, skip_label|None, warning|None)。"""
    business = clean_str(_cell(row, cols["买卖方向"]))
    if business in _SKIP_NAMES[fmt]:
        return None, business, None
    if business in _BUY_NAMES[fmt]:
        side = "buy"
    elif business in _SELL_NAMES[fmt]:
        side = "sell"
    else:
        return None, None, f"未知业务名称「{business}」,已跳过该行({ref})。"

    trade_date = _parse_cell_date(_cell(row, cols["成交日期"]))
    if trade_date is None:
        return None, None, f"交易日期无法解析,已跳过该行({ref})。"

    qty_raw = _parse_cell_float(_cell(row, cols["成交数量"]))
    if not qty_raw:
        return None, None, f"成交数量缺失或为 0,已跳过该行({ref})。"
    qty = int(round(abs(qty_raw)))

    name = clean_str(_cell(row, cols["名称"]))
    cash_flow = _parse_cell_float(_cell(row, cols["成交金额"])) or 0.0

    if fmt == "format1":
        # 格式一价格靠 (|发生金额|-费用)/数量 反推——这两列若压根没找到实际列
        # (默认列名/col_map 均未命中,不是"该行这两格恰好留空"),继续按 0 兜底
        # 会悄悄算出一个错误的价格(§4D.1 施工期实测踩过),必须当硬性缺列处理,
        # 不能像格式二那些"可能确实为 0"的杂费列一样静默兜底。
        if cols["成交金额"] is None:
            return None, None, f"未找到「发生金额」列(默认列名/col_map 均未命中),反推成交价需要该列,已跳过该行({ref})。"
        if cols["手续费"] is None:
            return None, None, f"未找到「手续费」对应列(默认列名「费用」/col_map 均未命中),反推成交价需要该列,已跳过该行({ref})。"
        fee_raw = _parse_cell_float(_cell(row, cols["手续费"])) or 0.0
        price = (abs(cash_flow) - fee_raw) / qty if qty else 0.0
        fee = abs(fee_raw)
        code = name_index.resolve(name, trade_date)
        if code is None:
            return None, None, f"证券名称「{name}」无法反查到代码(歧义或未匹配),已跳过该行({ref})。"
    else:
        raw_code = clean_str(_cell(row, cols["代码"]))
        if not raw_code:
            return None, None, f"证券代码缺失,已跳过该行({ref})。"
        code = normalize_ts_code(raw_code)
        price_cell = _parse_cell_float(_cell(row, cols["成交价"]))
        if price_cell is None:
            return None, None, f"成交价格缺失,已跳过该行({ref})。"
        price = price_cell
        commission = abs(_parse_cell_float(_cell(row, cols["手续费"])) or 0.0)
        stamp_tax = abs(_parse_cell_float(_cell(row, cols["印花税"])) or 0.0)
        transfer_fee = abs(_parse_cell_float(_cell(row, cols["过户费"])) or 0.0)
        clearing_fee = abs(_parse_cell_float(_cell(row, clearing_fee_idx)) or 0.0)
        fee = commission + stamp_tax + transfer_fee + clearing_fee

    if price <= 0:
        return None, None, f"反推/读取的成交价 <= 0(异常数据),已跳过该行({ref})。"

    trade = RawTrade(
        trade_date=trade_date, ts_code=code, name=name, side=side,
        price=round(price, 4), qty=qty, fee=round(fee, 2), cash_flow=cash_flow,
        source_format=fmt, source_ref=ref,
    )
    return trade, None, None


__all__ = [
    "CANONICAL_FIELDS",
    "clean_str",
    "normalize_ts_code",
    "NameIndex",
    "build_name_index",
    "RawTrade",
    "ParseWarning",
    "ParseResult",
    "parse_workbook",
]
