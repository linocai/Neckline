"""交割单解析(plan 4D.1)。输入券商导出的 xlsx(bytes)→ 规范化 `RawTrade` 列表。

**两家券商 schema(用户样例文件已不在磁盘,以下结构从真实文件逐字段核实过,任务
指令原文钉死,不臆造)**:

    · 券商一格式(手机截图整理,无成交价,也【无证券代码列】):
      交易日期(datetime) / 券商来源 / 业务名称 / 证券名称 / 成交数量 / 股份余额 /
      费用 / 发生金额 / 资金余额 / 备注。业务名称枚举:证券买入清算、证券卖出清算、
      银证转入、银证转出、红股派息、股息红利个人所得税扣款、利息归本。
      成交价反推:`(|发生金额| - 费用) / 数量`(公式按任务指令原样实现,不反向验证
      买卖方向下的正负号语义——发生金额/费用的符号约定由该券商决定,本模块不猜)。
      **无代码列 → ts_code 必须靠"证券名称"反查**(`resolve_code_by_name`)。

    · 券商二格式(桌面导出):交易日期 / 券商来源 / 证券代码 / 证券名称 / 业务名称 /
      成交价格 / 成交数量 / 成交金额 / 佣金 / 印花税 / 过户费 / 清算费(B股) /
      发生金额 / 资金余额 / 股份余额 / 股东代码 / 币种 / 合同编号。业务名称枚举:
      证券买入、证券卖出、银行转证券、指定交易。**证券代码等字段可能带零宽空格
      (U+200B)前缀,解析时 strip**;代码无交易所后缀,按 6/0/3/8-4-920 前缀补
      .SH/.SZ/.BJ(复用 `neckline.data.realtime.to_symbol`,它本身已复用
      `neckline.data.board.classify_by_code` 单一源,本模块不再另写一份前缀正则)。

通用清洗:跳过非交易行(指定交易 / 银证转账等资金流水,非成交)、金额正负号由
该券商自行约定(反推公式照抄任务指令,不做符号假设)、按(代码,日期序)供上层
FIFO 闭合回合(本模块只产出规范化交易流,不做回合闭合——见 `reconcile.py`)。

**可配字段映射(plan 4D.1「留 review_col_map 可覆盖以支持两家券商原始格式」)**:
`col_map` 是「规范字段 → 该工作簿实际列名」的覆盖字典,规范字段固定为 9 个
(`CANONICAL_FIELDS`);未覆盖的字段退化到 `_DEFAULT_HEADERS` 的内置默认列名。
表头行探测固定按"含《交易日期》列"(任务指令原文),不受 col_map 影响——col_map
只remap 其它字段的列名,不改表头探测锚点。
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional

import polars as pl

from neckline.data.market_data import load_namechange, load_stock_basic
from neckline.data.realtime import to_symbol

# —— 规范字段(plan 4D.1 原文枚举,col_map 的 key 空间)——————————————————————
CANONICAL_FIELDS = (
    "成交日期", "代码", "名称", "买卖方向", "成交价",
    "成交数量", "成交金额", "手续费", "印花税", "过户费",
)

# —— 业务名称枚举(两家券商固定值,非 col_map 可配范围——col_map 只 remap 列名,
#    不 remap 列值语义)——————————————————————————————————————————————————
_BUY_NAMES = {"format1": {"证券买入清算"}, "format2": {"证券买入"}}
_SELL_NAMES = {"format1": {"证券卖出清算"}, "format2": {"证券卖出"}}
_SKIP_NAMES = {
    "format1": {"银证转入", "银证转出", "红股派息", "股息红利个人所得税扣款", "利息归本"},
    "format2": {"银行转证券", "指定交易"},
}

# —— 资金流水四分类(plan §五 V2-⑫-A 扩展,2026-08-03)————————————————————————
# 蓝图 5.3 原文:"资金转入、转出、分红、税费和交易盈亏必须分开;账户金额增加不得
# 直接视为策略收益"。**这四类里只有前三类(转入转出/分红/税费)出自本模块的"跳过
# 行"**——"交易盈亏"来自 `reconcile.py` 的 FIFO 回合净盈亏(`RoundTrip.net_pnl`),
# 不是一种"业务名称";它的分类常量 `CASH_FLOW_TRADING_PNL` 因此不在本模块声明,
# 与这三类一起在更上层的 `review/cashflow.py` 统一导出,调用方按同一份四分类
# 常量拼装周度资金流水摘要,不在两处各写一份。
#
# ⚠ **对账引擎既有行为原样不动**:以下全部是新增能力,`_parse_row`/`_SKIP_NAMES`/
# `parse_workbook` 对成交行(买/卖)的既有解析路径一字未改;跳过行原有的
# `skip_counts` 计数行为也不变,本次只是**额外**把这些行的日期/金额/分类留痕到
# 新增的 `ParseResult.cash_flow_events`,供 `review/cashflow.py` 消费。
CASH_FLOW_TRANSFER = "transfer"   # 转入转出:银证转入/转出(格式一)、银行转证券(格式二)——符号即方向
CASH_FLOW_DIVIDEND = "dividend"   # 分红:红股派息等公司分配
CASH_FLOW_TAX = "tax"             # 税费:股息红利个人所得税扣款等
CASH_FLOW_OTHER = "other"         # 不属前三类的资金流水事件(如"利息归本"——broker 结算账户利息,
                                   # 既非转账也非分红/税费)或零金额账户操作(如"指定交易"——纯登记
                                   # 动作,不是资金流水)。**如实标"其他",不强并入分红/税费**——
                                   # 把利息记成分红、把账户操作记成税费都是伪造分类。

# 「业务名称 → 资金流水分类」唯一映射(两家券商跳过名合并成一张表——两边字面量
# 互不重叠,可以安全合并;新增券商格式时如果引入新的跳过名,必须同步在此登记,
# 否则会静默落 `CASH_FLOW_OTHER` 兜底——**兜底是故意的**,不识别的业务名称宁可
# 如实标"其他"也不可判"没有这笔事件")。
_CASH_FLOW_KIND_BY_NAME: Dict[str, str] = {
    "银证转入": CASH_FLOW_TRANSFER,
    "银证转出": CASH_FLOW_TRANSFER,
    "银行转证券": CASH_FLOW_TRANSFER,
    "红股派息": CASH_FLOW_DIVIDEND,
    "股息红利个人所得税扣款": CASH_FLOW_TAX,
    "利息归本": CASH_FLOW_OTHER,
    "指定交易": CASH_FLOW_OTHER,
}

# 「发生金额」列名(⚠ 不入 col_map——它不是任务指令原文枚举的 9 个规范字段之一,
# 同 `_FORMAT2_CLEARING_FEE_HEADER` 的既定体例:固定按此列名探测,两家格式的表头
# 都确有这一列〔见模块头两份 schema 描述〕,找不到该列时 ⛔ 不猜是哪一列,如实
# 放弃该行的资金流水留痕〔不影响 `skip_counts` 计数,那是另一件事〕)。
_CASH_FLOW_AMOUNT_HEADER = "发生金额"

# —— 内置默认列名(未被 col_map 覆盖时的兜底;§4D.1"先支持整理格式,留可配支持
#    两家券商原始格式")—————————————————————————————————————————————————
_DEFAULT_HEADERS = {
    "format1": {
        "成交日期": "交易日期",
        "名称": "证券名称",
        "买卖方向": "业务名称",
        "成交数量": "成交数量",
        "成交金额": "发生金额",   # 格式一无独立"成交金额"列,兜底用发生金额(反推价格用)
        "手续费": "费用",
    },
    "format2": {
        "成交日期": "交易日期",
        "代码": "证券代码",
        "名称": "证券名称",
        "买卖方向": "业务名称",
        "成交价": "成交价格",
        "成交数量": "成交数量",
        "成交金额": "成交金额",
        "手续费": "佣金",
        "印花税": "印花税",
        "过户费": "过户费",
    },
}
_FORMAT2_CLEARING_FEE_HEADER = "清算费(B股)"   # 不入 col_map(非任务指令枚举的 9 个规范字段之一),固定按此列名探测,缺失视为 0

_HEADER_SCAN_ROWS = 15   # 表头前最多容许 1-3 行标题/说明,留宽松缓冲

# 零宽字符(U+200B 零宽空格为任务指令明确提到的坑;顺带清一批同类不可见字符防御)。
_INVISIBLE_CHARS_RE = re.compile("[​‌‍﻿]")


def clean_str(v: object) -> str:
    """去首尾空白 + 零宽字符(§4D.1「证券代码等字段可能带零宽空格前缀」)。"""
    if v is None:
        return ""
    s = str(v)
    s = _INVISIBLE_CHARS_RE.sub("", s)
    s = unicodedata.normalize("NFKC", s)
    return s.strip()


def normalize_ts_code(raw: str) -> str:
    """裸 6 位代码(可能带零宽空格)→ 带交易所后缀的 ts_code。**不写正则**——
    复用 `neckline.data.realtime.to_symbol`(已复用 `board.classify_by_code`
    单一源判定北交所前缀,§4D.1「勿另写正则」)。已带 `.SH/.SZ/.BJ` 后缀的原样
    规范化大写返回。"""
    s = clean_str(raw)
    if not s:
        return s
    if re.match(r"^\d{6}\.[A-Za-z]{2}$", s):
        digits, suffix = s.split(".")
        return f"{digits}.{suffix.upper()}"
    digits = re.sub(r"\D", "", s)
    if not digits:
        return s
    sym = to_symbol(digits)          # 如 "sh600519" / "sz000001" / "bj920117"
    prefix, code_digits = sym[:2], sym[2:]
    suffix = {"sh": "SH", "sz": "SZ", "bj": "BJ"}.get(prefix, "SH")
    return f"{code_digits}.{suffix}"


def _parse_cell_date(v: object) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = clean_str(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_cell_time(v: object) -> Optional[time]:
    """成交**时刻**(北京时间,可选)。两家已知券商格式的「交易日期」列都只到日期粒度
    (格式一是 datetime 但时分秒为 0),故本函数绝大多数情况返回 `None` —— 那是正常的,
    不是解析失败:`reconcile.trade_instant` 会按「该日收盘时刻」兜底(v1.4-⑥-A 定死口径)。

    只在**确实带非零时刻**时才返回时刻:datetime 单元格取其 time(全零视作"只有日期",
    返 None,免得把 00:00 当成"凌晨成交"——A 股不存在的时刻);字符串按
    `'YYYY-MM-DD HH:MM(:SS)'` 两种格式试解。**绝不猜**:解析不出就 None。"""
    if isinstance(v, datetime):
        t = v.time()
        return t if (t.hour or t.minute or t.second) else None
    if isinstance(v, date):
        return None
    s = clean_str(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            t = datetime.strptime(s, fmt).time()
        except ValueError:
            continue
        return t if (t.hour or t.minute or t.second) else None
    return None


def _parse_cell_float(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = clean_str(v).replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


@dataclass
class NameIndex:
    """"证券名称 → ts_code" 反查索引(供格式一——无代码列——用)。按 trade_date
    做 as-of 区间匹配,`namechange` 覆盖不到的(已知有分页缺行小概率,阶段0 遗留
    问题 4)回退 `stock_basic` 当前名称精确匹配。同名映射到 >1 个不同 ts_code
    (歧义)时诚实返回 None,不瞎猜。"""

    namechange: pl.DataFrame
    stock_basic: pl.DataFrame

    def resolve(self, name: str, trade_date: date) -> Optional[str]:
        name = clean_str(name)
        if not name:
            return None
        codes: set = set()
        if not self.namechange.is_empty():
            sub = self.namechange.filter(pl.col("name") == name)
            for row in sub.iter_rows(named=True):
                start = row.get("start_date")
                end = row.get("end_date")
                if start is not None and start > trade_date:
                    continue
                if end is not None and end < trade_date:
                    continue
                codes.add(row["ts_code"])
        if len(codes) == 1:
            return next(iter(codes))
        if len(codes) > 1:
            return None  # 歧义:同名在同一时点映射到多个代码,不猜
        if not self.stock_basic.is_empty():
            sub2 = self.stock_basic.filter(pl.col("name") == name)
            codes2 = set(sub2["ts_code"].to_list())
            if len(codes2) == 1:
                return next(iter(codes2))
        return None


def build_name_index(db_path: Optional[Path] = None) -> NameIndex:
    return NameIndex(namechange=load_namechange(db_path), stock_basic=load_stock_basic(db_path))


@dataclass
class RawTrade:
    trade_date: date
    ts_code: str
    name: str
    side: str                 # "buy" | "sell"
    price: float               # 成交价(格式一反推 / 格式二直读)
    qty: int                   # 成交数量(股,正数)
    fee: float                 # 本笔归属费用(绝对值;格式一=|费用|,格式二=佣金+印花税+过户费+清算费(B股) 各自取绝对值求和)
    cash_flow: float            # 发生金额(原始符号,供审计留痕,不参与三查计算)
    broker_source: str = ""     # 券商来源(如有)
    source_format: str = ""     # "format1" | "format2"
    source_ref: str = ""        # "<sheet>!row<N>",诊断用
    # v1.4-⑥-A:成交**时刻**(北京时间),仅当交割单确实带非零时刻时非空。两家已知格式
    # 都只到日期粒度 → 恒 None,由 `reconcile.trade_instant` 按「该日收盘时刻」兜底判章程。
    trade_time: Optional[time] = None


@dataclass
class ParseWarning:
    sheet: str
    row: int
    message: str


@dataclass
class CashFlowEvent:
    """⑫-A 扩展:一条非成交行(`_SKIP_NAMES`命中)的资金流水留痕。**只读派生**——
    不改变 `skip_counts` 既有计数语义,是同一批跳过行的另一份视图(带日期/金额/
    四分类,供 `review/cashflow.py` 按周聚合)。"""

    trade_date: date
    business_name: str        # 原始业务名称(留痕,如"红股派息")
    kind: str                 # CASH_FLOW_* 之一
    amount: float              # 「发生金额」原始符号(转入转出的方向即由符号表达)
    ts_code: Optional[str] = None   # 关联证券(格式二的代码列直读;格式一账户级事件无代码 → None)
    name: str = ""             # 证券名称(账户级事件为空串,如银证转入/转出)
    source_format: str = ""    # "format1" | "format2"
    source_ref: str = ""       # "<sheet>!row<N>",诊断用


@dataclass
class ParseResult:
    trades: List[RawTrade] = field(default_factory=list)
    warnings: List[ParseWarning] = field(default_factory=list)
    sheet_formats: Dict[str, str] = field(default_factory=dict)   # sheet名 -> "format1"/"format2"/"skipped:<reason>"
    skip_counts: Dict[str, int] = field(default_factory=dict)     # 业务名称(非成交行)-> 跳过行数,供展示留痕
    cash_flow_events: List[CashFlowEvent] = field(default_factory=list)   # ⑫-A:跳过行的资金流水留痕(新增,不影响以上既有字段)


def _find_header_row(rows: List[List[object]]) -> Optional[int]:
    """在前 `_HEADER_SCAN_ROWS` 行内找"含《交易日期》列"的那一行(任务指令原文
    锚点,不受 col_map 影响)。返回该行在 `rows` 中的下标,找不到 → None。"""
    limit = min(len(rows), _HEADER_SCAN_ROWS)
    for i in range(limit):
        cells = [clean_str(c) for c in rows[i]]
        if "交易日期" in cells:
            return i
    return None


def _detect_format(headers: List[str], col_map: Optional[Dict[str, str]] = None) -> Optional[str]:
    """格式二有独立"证券代码"列(格式一没有,只有"证券名称")——用它作主判据;
    "成交价格"列是次要判据(格式一无成交价,靠反推)。两者皆无 → 判格式一
    (格式一字段集是格式二的子集去掉"证券代码"/"成交价格"等,故默认落这里)。

    判据列名本身也吃 col_map 覆盖(若用户的券商格式把"证券代码"/"成交价格"/
    "证券名称"这几个判据列改了名,不 remap 判据就永远探测不出格式,col_map 就
    形同虚设)——表头行探测锚点("交易日期")仍固定不受影响,这里只是"认出
    是哪种格式"的判据列名可配。"""
    m = col_map or {}
    code_col = m.get("代码", "证券代码")
    price_col = m.get("成交价", "成交价格")
    name_col = m.get("名称", "证券名称")
    if code_col in headers or price_col in headers:
        return "format2"
    if name_col in headers:
        return "format1"
    return None


def _col_index(headers: List[str], canonical: str, fmt: str, col_map: Optional[Dict[str, str]]) -> Optional[int]:
    override = (col_map or {}).get(canonical)
    target = override if override else _DEFAULT_HEADERS[fmt].get(canonical)
    if not target:
        return None
    try:
        return headers.index(target)
    except ValueError:
        return None


def _cell(row: List[object], idx: Optional[int]) -> object:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _extract_cash_flow_event(
    row: List[object],
    business: str,
    fmt: str,
    cols: Dict[str, Optional[int]],
    cash_flow_idx: Optional[int],
    ref: str,
) -> Optional[CashFlowEvent]:
    """⑫-A 扩展:把一条已判定为「跳过」的非成交行(`business in _SKIP_NAMES[fmt]`)
    一并留痕成 `CashFlowEvent`。**只在这里新增能力,不改 `_parse_row` 一个字节**
    ——调用方(`parse_workbook`)在拿到 `skip_label` 之后另行调用本函数,两条路径
    互不干扰。

    日期解不出 / 找不到「发生金额」列(两家已知格式表头都确有此列,找不到多半是
    col_map 覆盖了非本函数管辖的字段名)→ 如实返回 `None`(该行仍计入既有
    `skip_counts`,只是资金流水这一份"新视图"缺席,**绝不**猜一个日期或金额凑数)。
    **本函数永不抛异常、永不产生 `ParseWarning`**——已知的跳过行本就不该有警告噪音
    (`test_format1_skips_non_trade_rows` 断言 `warnings == []`,四分类留痕不能打破
    这条既有断言)。"""
    trade_date = _parse_cell_date(_cell(row, cols["成交日期"]))
    if trade_date is None or cash_flow_idx is None:
        return None
    amount = _parse_cell_float(_cell(row, cash_flow_idx))
    if amount is None:
        amount = 0.0
    name = clean_str(_cell(row, cols["名称"]))
    ts_code: Optional[str] = None
    if fmt == "format2":
        # 格式二的「证券代码」列对跳过行(银行转证券/指定交易)同样直读——这两类
        # 业务名称本身就与某只证券的账户操作相关(见模块头两家 schema 描述),不需要
        # 反查;格式一的账户级事件(银证转入/转出/利息归本)没有代码列,`ts_code`
        # 恒 None。**红股派息/股息红利个人所得税扣款虽有「证券名称」,本函数不额外
        # 反查代码**(需要 `NameIndex`,本函数刻意不接它以保持零依赖/零警告)——
        # `name` 字段已留痕,够审计用,不为了补全一个 ts_code 而引入新的失败模式。
        raw_code = clean_str(_cell(row, cols["代码"]))
        ts_code = normalize_ts_code(raw_code) if raw_code else None
    return CashFlowEvent(
        trade_date=trade_date, business_name=business,
        kind=_CASH_FLOW_KIND_BY_NAME.get(business, CASH_FLOW_OTHER),
        amount=amount, ts_code=ts_code, name=name,
        source_format=fmt, source_ref=ref,
    )


def parse_workbook(
    data: bytes,
    filename: str = "",
    *,
    col_map: Optional[Dict[str, str]] = None,
    db_path: Optional[Path] = None,
    name_index: Optional[NameIndex] = None,
) -> ParseResult:
    """解析一份交割单 xlsx(多 sheet)。任何单行解析异常降级为该行的一条
    `ParseWarning`(跳过该行,不中断整份解析);整份工作簿打不开(非法 xlsx)→
    抛 `ValueError`,由 API 层转 400(这是"文件本身不是 xlsx"的用户可纠正错误,
    与"某一行数据有瑕疵"的优雅降级是两回事)。"""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"{filename or '文件'} 不是合法的 xlsx 工作簿:{e}") from e

    result = ParseResult()
    idx = name_index if name_index is not None else build_name_index(db_path)

    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            result.sheet_formats[ws.title] = "skipped:empty"
            continue
        header_i = _find_header_row(rows)
        if header_i is None:
            result.sheet_formats[ws.title] = "skipped:no_header(未找到含'交易日期'的表头行)"
            continue
        headers = [clean_str(c) for c in rows[header_i]]
        fmt = _detect_format(headers, col_map)
        if fmt is None:
            result.sheet_formats[ws.title] = "skipped:unknown_format"
            result.warnings.append(ParseWarning(ws.title, header_i + 1, "表头列不匹配已知的两种券商格式,已跳过整个 sheet。"))
            continue
        result.sheet_formats[ws.title] = fmt

        cols = {c: _col_index(headers, c, fmt, col_map) for c in CANONICAL_FIELDS}
        clearing_fee_idx = headers.index(_FORMAT2_CLEARING_FEE_HEADER) if _FORMAT2_CLEARING_FEE_HEADER in headers else None
        # ⑫-A 扩展:「发生金额」不入 col_map(同 `clearing_fee_idx` 体例),固定列名探测。
        cash_flow_idx = headers.index(_CASH_FLOW_AMOUNT_HEADER) if _CASH_FLOW_AMOUNT_HEADER in headers else None

        for r in range(header_i + 1, len(rows)):
            row = rows[r]
            if row is None or all(c is None or clean_str(c) == "" for c in row):
                continue
            ref = f"{ws.title}!row{r + 1}"
            try:
                trade, skip_label, warn_msg = _parse_row(
                    row, fmt, cols, clearing_fee_idx, idx, ref,
                )
            except Exception as e:  # noqa: BLE001  单行异常绝不中断整份解析
                result.warnings.append(ParseWarning(ws.title, r + 1, f"解析异常已跳过该行:{e}"))
                continue
            if warn_msg:
                result.warnings.append(ParseWarning(ws.title, r + 1, warn_msg))
            if skip_label:
                result.skip_counts[skip_label] = result.skip_counts.get(skip_label, 0) + 1
                try:  # ⑫-A 扩展:资金流水留痕是可选的"新视图",失败绝不牵连既有解析
                    event = _extract_cash_flow_event(row, skip_label, fmt, cols, cash_flow_idx, ref)
                except Exception:  # noqa: BLE001
                    event = None
                if event is not None:
                    result.cash_flow_events.append(event)
            if trade is not None:
                result.trades.append(trade)

    return result


def _parse_row(
    row: List[object],
    fmt: str,
    cols: Dict[str, Optional[int]],
    clearing_fee_idx: Optional[int],
    name_index: NameIndex,
    ref: str,
):
    """返回 (RawTrade|None, skip_label|None, warning|None)。"""
    business = clean_str(_cell(row, cols["买卖方向"]))
    if business in _SKIP_NAMES[fmt]:
        return None, business, None
    if business in _BUY_NAMES[fmt]:
        side = "buy"
    elif business in _SELL_NAMES[fmt]:
        side = "sell"
    else:
        return None, None, f"未知业务名称「{business}」,已跳过该行({ref})。"

    date_cell = _cell(row, cols["成交日期"])
    trade_date = _parse_cell_date(date_cell)
    if trade_date is None:
        return None, None, f"交易日期无法解析,已跳过该行({ref})。"
    trade_time = _parse_cell_time(date_cell)   # v1.4-⑥-A:带时刻就取,不带就 None(不猜)

    qty_raw = _parse_cell_float(_cell(row, cols["成交数量"]))
    if not qty_raw:
        return None, None, f"成交数量缺失或为 0,已跳过该行({ref})。"
    qty = int(round(abs(qty_raw)))

    name = clean_str(_cell(row, cols["名称"]))
    cash_flow = _parse_cell_float(_cell(row, cols["成交金额"])) or 0.0

    if fmt == "format1":
        # 格式一价格靠 (|发生金额|-费用)/数量 反推——这两列若压根没找到实际列
        # (默认列名/col_map 均未命中,不是"该行这两格恰好留空"),继续按 0 兜底
        # 会悄悄算出一个错误的价格(§4D.1 施工期实测踩过),必须当硬性缺列处理,
        # 不能像格式二那些"可能确实为 0"的杂费列一样静默兜底。
        if cols["成交金额"] is None:
            return None, None, f"未找到「发生金额」列(默认列名/col_map 均未命中),反推成交价需要该列,已跳过该行({ref})。"
        if cols["手续费"] is None:
            return None, None, f"未找到「手续费」对应列(默认列名「费用」/col_map 均未命中),反推成交价需要该列,已跳过该行({ref})。"
        fee_raw = _parse_cell_float(_cell(row, cols["手续费"])) or 0.0
        price = (abs(cash_flow) - fee_raw) / qty if qty else 0.0
        fee = abs(fee_raw)
        code = name_index.resolve(name, trade_date)
        if code is None:
            return None, None, f"证券名称「{name}」无法反查到代码(歧义或未匹配),已跳过该行({ref})。"
    else:
        raw_code = clean_str(_cell(row, cols["代码"]))
        if not raw_code:
            return None, None, f"证券代码缺失,已跳过该行({ref})。"
        code = normalize_ts_code(raw_code)
        price_cell = _parse_cell_float(_cell(row, cols["成交价"]))
        if price_cell is None:
            return None, None, f"成交价格缺失,已跳过该行({ref})。"
        price = price_cell
        commission = abs(_parse_cell_float(_cell(row, cols["手续费"])) or 0.0)
        stamp_tax = abs(_parse_cell_float(_cell(row, cols["印花税"])) or 0.0)
        transfer_fee = abs(_parse_cell_float(_cell(row, cols["过户费"])) or 0.0)
        clearing_fee = abs(_parse_cell_float(_cell(row, clearing_fee_idx)) or 0.0)
        fee = commission + stamp_tax + transfer_fee + clearing_fee

    if price <= 0:
        return None, None, f"反推/读取的成交价 <= 0(异常数据),已跳过该行({ref})。"

    trade = RawTrade(
        trade_date=trade_date, ts_code=code, name=name, side=side,
        price=round(price, 4), qty=qty, fee=round(fee, 2), cash_flow=cash_flow,
        source_format=fmt, source_ref=ref, trade_time=trade_time,
    )
    return trade, None, None


__all__ = [
    "CANONICAL_FIELDS",
    "clean_str",
    "normalize_ts_code",
    "NameIndex",
    "build_name_index",
    "RawTrade",
    "ParseWarning",
    "ParseResult",
    "parse_workbook",
    # ⑫-A 资金流水四分类扩展
    "CASH_FLOW_TRANSFER",
    "CASH_FLOW_DIVIDEND",
    "CASH_FLOW_TAX",
    "CASH_FLOW_OTHER",
    "CashFlowEvent",
]
